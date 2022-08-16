import openpyxl

book = openpyxl.load_workbook("C:\\Users\\PrafulS\\Downloads\\ExecelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=2,column=3)
print(cell.value)
Dict={}
#write operation
sheet.cell(row=3, column=2).value="Akash"
print(sheet.cell(row=3, column=2).value)

#how to know row column count
print(sheet.max_row)
print(sheet.max_column)

print(sheet['A3'].value)

for i in range(1,sheet.max_row+1):
    if(sheet.cell(row=i,column=1).value)=="Testcase2":
        for j in range(1,sheet.max_column+1):
            #print(sheet.cell(row=i, column=j).value)

            #load excel data into dictionary
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
            print(Dict)
