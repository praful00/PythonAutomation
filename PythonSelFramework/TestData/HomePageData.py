import openpyxl


class Homepagedata:

    test_Homepage_data=[{"firstname":"Praful","email":"praful@gmail.com","password":"12345","gender":"Female"},{"firstname":"Rahul","email":"rahul@gmail.com","password":"12346","gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\PrafulS\\Downloads\\ExecelDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if (sheet.cell(row=i, column=1).value) == "Testcase2":
                for j in range(1, sheet.max_column + 1):
                    # print(sheet.cell(row=i, column=j).value)

                    # load excel data into dictionary
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]