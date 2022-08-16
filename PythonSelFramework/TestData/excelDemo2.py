import openpyxl

book = openpyxl.load_workbook("C:\\Users\\PrafulS\\Downloads\\ExcelDemo2.xlsx")
sheet = book.active
cell= sheet.cell(row=1, column=1)
print(cell.value)
Dict={}
print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row+1):
    #if (sheet.cell(row=i,column=1).value) == 4:
        for j in range(1,sheet.max_column):
            print(sheet.cell(row=i,column=j).value)


#printing all data fro excel
